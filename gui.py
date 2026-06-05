import os
import sys
import json
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from threading import Thread
import time
import re
import threading
import logging
import shutil

# 版本号
VERSION = "v3.1"

# 自动更新配置
VERSION_FILE = "version.json"
UPDATE_EXE_NAME = "药品批发企业追朔码自动处理软件.exe"

if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
    # 安装模式下可写文件重定向到 %APPDATA%
    DATA_DIR = os.path.join(os.environ.get('APPDATA', os.path.expanduser("~")), "药品批发企业追朔码自动处理软件")
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    DATA_DIR = BASE_DIR
os.makedirs(DATA_DIR, exist_ok=True)

def compare_versions(v1, v2):
    """比较两个版本号，返回1表示v1>v2，-1表示v1<v2，0表示相等"""
    def normalize(v):
        # 移除'v'前缀，转为数字列表
        v = v.lstrip('vV')
        parts = []
        for part in re.split(r'[.\-_]', v):
            # 提取数字部分
            num = ''
            for c in part:
                if c.isdigit():
                    num += c
                else:
                    break
            parts.append(int(num) if num else 0)
        return parts

    p1, p2 = normalize(v1), normalize(v2)
    # 补齐长度
    while len(p1) < len(p2): p1.append(0)
    while len(p2) < len(p1): p2.append(0)

    for a, b in zip(p1, p2):
        if a > b: return 1
        if a < b: return -1
    return 0

def check_and_update():
    """检查更新：如有新版本则自动更新后重启"""
    version_file = os.path.join(BASE_DIR, VERSION_FILE)
    current_exe = sys.executable
    
    # 检查版本文件是否存在
    if not os.path.exists(version_file):
        return False
    
    try:
        with open(version_file, 'r', encoding='utf-8') as f:
            update_info = json.load(f)
        latest_version = update_info.get('version', '')
        
        if not latest_version:
            return False
        
        # 比较版本：如果最新版本 > 当前版本
        if compare_versions(latest_version, VERSION) > 0:
            print(f"[更新] 发现新版本 {latest_version}，当前版本 {VERSION}，正在更新...")
            
            # 复制新exe覆盖当前exe
            new_exe = os.path.join(BASE_DIR, UPDATE_EXE_NAME)
            if os.path.exists(new_exe):
                # 先删除旧的备份文件（如果存在）
                backup_exe = current_exe + ".old"
                if os.path.exists(backup_exe):
                    try:
                        os.remove(backup_exe)
                    except:
                        pass
                
                # 先把当前exe重命名为.bak，避免覆盖失败
                try:
                    os.rename(current_exe, backup_exe)
                except:
                    pass
                
                # 复制新exe
                shutil.copy2(new_exe, current_exe)
                print("[更新] 更新完成，正在启动新版本...")
                
                # 启动新版本后退出
                os.startfile(current_exe)
                return True  # 需要退出
            else:
                print(f"[更新] 新版本exe不存在: {new_exe}")
                
    except Exception as e:
        print(f"[更新] 检查更新失败: {e}")
    
    return False

CONFIG_FILE = os.path.join(DATA_DIR, 'config.json')
LOG_FILE = os.path.join(DATA_DIR, 'debug.log')
OUTPUT_FOLDER = r'E:\desktop\各大医院追溯码'
DEFAULT_INPUT_PATH = r'E:\desktop\各大医院追溯码'
# 如果输出目录不存在，回退到桌面
if not os.path.exists(OUTPUT_FOLDER):
    OUTPUT_FOLDER = os.path.join(os.path.expanduser("~"), "Desktop", "追溯码输出")
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
if not os.path.exists(DEFAULT_INPUT_PATH):
    DEFAULT_INPUT_PATH = os.path.expanduser("~\\Desktop")  # 回退到桌面

# 配置日志 - 同时输出到文件
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8', mode='w'),
    ]
)
logger = logging.getLogger(__name__)

# ============================================================
# 数据模型
# ============================================================
from dataclasses import dataclass, field

@dataclass
class TraceRecord:
    trace_code: str
    drug_name: str = ""
    dosage_form: str = ""
    pack_spec: str = ""
    approval_no: str = ""
    manufacturer: str = ""
    code_type: str = ""
    code_pack_count: str = ""
    loose_count: str = ""
    production_info: str = ""
    batch_no: str = ""
    verify_info: str = ""
    scan_time: str = ""

@dataclass
class MetaInfo:
    scan_date: str = ""
    doc_type: str = ""
    doc_no: str = ""
    receiver_name: str = ""

@dataclass
class ExcelData:
    meta: MetaInfo = field(default_factory=MetaInfo)
    headers: list = field(default_factory=list)
    records: list = field(default_factory=list)

# ============================================================
# Excel 读写（openpyxl 延迟加载，加速启动）
# ============================================================

def _safe_str(val) -> str:
    if val is None: return ""
    return str(val).strip()

def read_sales_excel(file_path: str) -> ExcelData:
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    data = ExcelData()
    for row_idx in range(1, 6):
        cell_a = ws.cell(row=row_idx, column=1).value
        cell_b = ws.cell(row=row_idx, column=2).value
        key, value = _safe_str(cell_a), _safe_str(cell_b)
        if key == "扫码日期": data.meta.scan_date = value
        elif key == "单据类型": data.meta.doc_type = value
        elif key == "单据号": data.meta.doc_no = value
        elif key == "收货单位名称": data.meta.receiver_name = value
    header_row = 6
    headers = []
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        headers.append(val if val else "")
    data.headers = headers
    
    # 按表头名称查找列索引（兼容不同列顺序）
    col_map = {}
    for idx, h in enumerate(headers):
        col_map[h] = idx
    
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        if not any(row_values): continue
        tc = row_values[0]
        if tc is None: continue
        tc = str(tc).strip()
        if not tc: continue
        # 空白码包装数处理：重读1次确认，仍空白则标记为'?'（异常码）
        raw_pack = row_values[col_map.get("码包装数", 7)]
        pack_str = _safe_str(raw_pack)
        if not pack_str:
            # 重读一次
            retry = ws.cell(row=row_idx, column=col_map.get("码包装数", 7) + 1).value
            if retry is not None and str(retry).strip():
                pack_str = str(retry).strip()
            else:
                pack_str = "?"  # 标记异常
                logger.debug(f"第{row_idx}行码包装数为空，标记为异常")
        
        data.records.append(TraceRecord(
            trace_code=tc,
            drug_name=_safe_str(row_values[col_map.get("药品器械名称", 1)]),
            dosage_form=_safe_str(row_values[col_map.get("剂型规格", 2)]),
            pack_spec=_safe_str(row_values[col_map.get("包装规格", 3)]),
            approval_no=_safe_str(row_values[col_map.get("批准文号", 4)]),
            manufacturer=_safe_str(row_values[col_map.get("生产厂家", 5)]),
            code_type=_safe_str(row_values[col_map.get("码类型", 6)]),
            code_pack_count=pack_str,
            loose_count=_safe_str(row_values[col_map.get("零头数", 8)]),
            production_info=_safe_str(row_values[col_map.get("生产信息", 9)]),
            batch_no=_safe_str(row_values[col_map.get("产品批号", 10)]),
            verify_info=_safe_str(row_values[col_map.get("验证信息", 11)]),
            scan_time=_safe_str(row_values[col_map.get("扫描时间", 12)]),
        ))
    wb.close()
    return data

def generate_filename(meta, drug_names=None, unit_name="药品批发企业"):
    """生成文件名：单位名称-收货单位-药品名(去重&连接)-单据号"""
    receiver = meta.receiver_name or "未知单位"
    doc_no = meta.doc_no or "未知单据号"
    if drug_names:
        # 去重并保持顺序
        seen = set()
        unique_names = []
        for name in drug_names:
            if name and name not in seen:
                seen.add(name)
                unique_names.append(name)
        name_str = "&".join(unique_names) if unique_names else "药品"
    else:
        name_str = "药品"
    for ch in r'\/:*?"<>|':
        receiver, name_str, doc_no = receiver.replace(ch,""), name_str.replace(ch,""), doc_no.replace(ch,"")
    return f"{unit_name}-{receiver}-{name_str}-{doc_no}.xlsx"

def write_result_excel(original_data, level_one_map, batch_no_map, output_dir, unit_name="药品批发企业"):
    """新建表格，每行写完整药品信息，删除零头数/生产信息/验证信息列"""
    import openpyxl
    from openpyxl.styles import Font, Alignment
    
    os.makedirs(output_dir, exist_ok=True)
    
    # 收集所有不重复的药品名（用于文件名）
    all_drug_names = []
    seen_names = set()
    for record in original_data.records:
        if record.drug_name and record.drug_name not in seen_names:
            seen_names.add(record.drug_name)
            all_drug_names.append(record.drug_name)
    
    output_path = os.path.join(output_dir, generate_filename(original_data.meta, all_drug_names, unit_name))
    
    # 新的表头：删除零头数(9)、生产信息(10)、验证信息(12)
    new_headers = [
        "追溯码", "药品器械名称", "剂型规格", "包装规格",
        "批准文号", "生产厂家", "码类型", "码包装数",
        "产品批号", "扫描时间"
    ]
    
    # 等线字体
    DENGXIAN_FONT = Font(name="等线")
    DENGXIAN_BOLD = Font(name="等线", bold=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # 写入前4行元信息（等线字体）
    for i, (k, v) in enumerate([("扫码日期", original_data.meta.scan_date), ("单据类型", original_data.meta.doc_type),
                                  ("单据号", original_data.meta.doc_no), ("收货单位名称", original_data.meta.receiver_name)], 1):
        cell_k = ws.cell(row=i, column=1, value=k)
        cell_k.font = DENGXIAN_FONT
        cell_v = ws.cell(row=i, column=2, value=v)
        cell_v.font = DENGXIAN_FONT
    
    # 写入表头（第6行）- 等线加粗，居中
    for col_idx, header in enumerate(new_headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.font = DENGXIAN_BOLD
        cell.alignment = Alignment(horizontal="left")    
    # 写入数据行
    data_row = 7
    for record in original_data.records:
        codes = level_one_map.get(record.trace_code, [record.trace_code])
        # 获取产品批号（从码上放心查询结果中获取）
        batch_no = batch_no_map.get(record.trace_code, record.batch_no) if batch_no_map else record.batch_no
        for code in codes:
            row_data = [
                code,                    # 1: 追溯码
                record.drug_name,        # 2: 药品器械名称
                record.dosage_form,      # 3: 剂型规格
                record.pack_spec,        # 4: 包装规格
                record.approval_no,      # 5: 批准文号
                record.manufacturer,     # 6: 生产厂家
                record.code_type,        # 7: 码类型
                "1",                     # 8: 码包装数（固定为1）
                batch_no,                # 9: 产品批号
                record.scan_time,        # 10: 扫描时间（复用输入表格同药品的时间）
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=data_row, column=col_idx, value=val)
                cell.font = DENGXIAN_FONT
                cell.alignment = Alignment(horizontal="left", vertical="center")
                # 第1列：追溯码设为文本格式
                if col_idx == 1:
                    cell.number_format = "@"
            data_row += 1
    
    # 列宽自适应：根据每列最长内容计算宽度
    for col_idx in range(1, len(new_headers) + 1):
        max_len = len(str(new_headers[col_idx - 1]))
        for row_idx in range(7, data_row):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                # 中文字符宽度约为英文2倍
                val_len = sum(2 if ord(c) > 127 else 1 for c in str(val))
                if val_len > max_len:
                    max_len = val_len
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 4
    
    wb.save(output_path)
    wb.close()
    return output_path

# ============================================================
# 多级码过滤
# ============================================================
def filter_level_one(raw_text):
    """从复制结果中过滤出1级码 - 找到'1级码'标记后，取其后面所有纯数字追溯码行"""
    if not raw_text or not raw_text.strip(): 
        return []
    
    codes = []
    lines = raw_text.strip().split("\n")
    
    logger.debug(f"原始内容:\n{raw_text[:500]}...")
    
    # 找到"1级码"标记的位置，然后取后面所有纯数字行
    found_level_one = False
    for line in lines:
        line = line.strip()
        if not line: 
            continue
        
        # 检测是否是级别标记（如"1级码"、"2级码"、"3级码"等）
        if re.match(r"^\d级码", line):
            if re.match(r"^1级码", line):
                found_level_one = True
                logger.debug(f"找到1级码标记")
            else:
                # 遇到其他级别标记（2级码、3级码等），停止收集
                found_level_one = False
            continue
        
        # 如果已经找到1级码标记，收集后续的纯数字行
        if found_level_one and re.match(r"^\d{16,24}$", line):
            codes.append(line)
    
    logger.debug(f"过滤后1级码共 {len(codes)} 个: {codes[:5]}{'...' if len(codes) > 5 else ''}")
    return codes


def extract_all_codes(raw_text):
    """从复制结果中提取全部码（3级+2级+1级），返回 set"""
    if not raw_text or not raw_text.strip():
        return set()
    
    all_codes = set()
    lines = raw_text.strip().split("\n")
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        # 跳过级别标记（3级码、2级码、1级码等）
        if re.match(r"^\d级码", line):
            continue
        # 收集所有16-24位的纯数字追溯码
        if re.match(r"^\d{16,24}$", line):
            all_codes.add(line)
    
    logger.debug(f"提取全部码共 {len(all_codes)} 个")
    return all_codes

# ============================================================
# UI 自动化 - 使用 pyautogui 模拟鼠标键盘
# ============================================================
import pyautogui
import pyperclip

# 全局键盘监听（pynput） - 用于F1强制停止、F5全局校准
try:
    from pynput import keyboard as pynput_keyboard
    PYNPUT_AVAILABLE = True
except ImportError:
    PYNPUT_AVAILABLE = False
    print("[警告] pynput 未安装，全局快捷键(F1/F5)将不可用")

# 安全设置
pyautogui.PAUSE = 0.3
pyautogui.FAILSAFE = True


class ScreenCalibrator:
    """屏幕坐标校准器"""
    
    def __init__(self, config_file):
        self.config_file = config_file
        self.positions = self._load()
    
    def _load(self):
        if os.path.exists(self.config_file):
            with open(self.config_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}
    
    def _save(self):
        with open(self.config_file, 'w', encoding='utf-8') as f:
            json.dump(self.positions, f, ensure_ascii=False, indent=2)
    
    def get(self, name):
        return self.positions.get(name)
    
    def set(self, name, x, y):
        self.positions[name] = {"x": x, "y": y}
        self._save()
    
    def is_calibrated(self):
        return all(k in self.positions for k in ["input_box", "query_btn", "copy_btn", "batch_no_pos", "copy_btn_after"])


class CodeQueryUI:
    """使用pyautogui模拟鼠标键盘操作码上放心客户端"""
    
    def __init__(self, calibrator, stop_flag=None, fail_callback=None):
        self.cal = calibrator
        self._pc = pyperclip
        self.stop_flag = stop_flag if stop_flag else threading.Event()
        self.fail_callback = fail_callback  # 失败回调，用于弹窗报警
        if not self.cal.is_calibrated():
            raise RuntimeError("请先完成校准！点击'校准位置'按钮")
    
    def _activate_window(self):
        """尝试激活码上放心窗口"""
        try:
            windows = pyautogui.getWindowsWithTitle("码上放心")
            if windows:
                windows[0].activate()
                time.sleep(0.3)
                return True
        except:
            pass
        return False
    
    def _click(self, name):
        pos = self.cal.get(name)
        if not pos:
            raise RuntimeError(f"未校准: {name}")
        self._activate_window()  # 每步点击前激活窗口
        pyautogui.click(pos['x'], pos['y'])
    
    def _type_code(self, code):
        pos = self.cal.get("input_box")
        self._activate_window()  # 每次输入前激活窗口
        # 点击输入框
        pyautogui.click(pos['x'], pos['y'])
        time.sleep(0.3)
        # 全选并删除
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(0.1)
        # 用剪贴板粘贴（避免输入法问题）
        pyperclip.copy(code)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(0.3)
    
    def query_single(self, trace_code):
        """查询追溯码，获取1级码列表、全部码集合和产品批号（带重试机制）"""
        result = {"success": False, "code": trace_code, "level_one_codes": [], "all_codes": set(), "batch_no": "", "error": ""}
        
        max_retries = 3
        for attempt in range(1, max_retries + 1):
            try:
                logger.debug(f"开始处理追溯码: {trace_code} (第{attempt}次尝试)")
                self._type_code(trace_code)
                time.sleep(0.3)
                self._click("query_btn")
                time.sleep(2)
                self._click("copy_btn_after")
                time.sleep(0.5)
                
                # 剪贴板检查 + 重试
                raw_text = pyperclip.paste()
                if not raw_text or not raw_text.strip():
                    if attempt < max_retries:
                        logger.debug(f"复制结果为空，重试 ({attempt}/{max_retries})")
                        time.sleep(1)
                        continue
                    result["error"] = f"重试{max_retries}次后复制结果仍为空"
                    logger.debug(result["error"])
                    return result
                
                level_one = filter_level_one(raw_text)
                result["level_one_codes"] = level_one
                result["all_codes"] = extract_all_codes(raw_text)
                
                # 批号获取（内部重试2次）
                batch_no = ""
                batch_pos = self.cal.get("batch_no_pos")
                if batch_pos:
                    for _ in range(2):
                        try:
                            pyautogui.doubleClick(batch_pos['x'], batch_pos['y'])
                            time.sleep(0.3)
                            pyautogui.hotkey('ctrl', 'c')
                            time.sleep(0.3)
                            batch_no = self._pc.paste().strip()
                            if batch_no:
                                break
                        except:
                            pass
                result["batch_no"] = batch_no
                result["success"] = True
                logger.debug(f"完成，{len(level_one)}个1级码，批号: {batch_no}")
                return result
                
            except Exception as e:
                logger.error(f"处理追溯码 {trace_code} (第{attempt}次) 出错: {str(e)}")
                if attempt < max_retries:
                    time.sleep(1)
                    continue
                result["error"] = str(e)
                return result
        
        return result
    
    def query_batch_only(self, trace_code):
        """只查询并获取产品批号（用于已经是1级码的情况），带重试机制"""
        result = {"success": False, "code": trace_code, "level_one_codes": [trace_code], "batch_no": "", "error": ""}
        
        max_retries = 3
        for attempt in range(1, max_retries + 1):
            try:
                logger.debug(f"[批号查询] {trace_code} (第{attempt}次)")
                self._type_code(trace_code)
                time.sleep(0.3)
                self._click("query_btn")
                time.sleep(2)
                
                batch_no = ""
                batch_pos = self.cal.get("batch_no_pos")
                if batch_pos:
                    for _ in range(2):
                        try:
                            pyautogui.doubleClick(batch_pos['x'], batch_pos['y'])
                            time.sleep(0.3)
                            pyautogui.hotkey('ctrl', 'c')
                            time.sleep(0.3)
                            batch_no = self._pc.paste().strip()
                            if batch_no:
                                break
                        except:
                            pass
                
                if not batch_no and attempt < max_retries:
                    logger.debug(f"批号为空，重试 ({attempt}/{max_retries})")
                    time.sleep(1)
                    continue
                
                result["batch_no"] = batch_no
                result["success"] = True
                logger.debug(f"批号: {batch_no or '（空）'}")
                return result
                
            except Exception as e:
                logger.error(f"获取批号出错 (第{attempt}次): {str(e)}")
                if attempt < max_retries:
                    time.sleep(1)
                    continue
                result["error"] = str(e)
                return result
        
        return result

    def _drag_select_right_panel(self):
        """
        在右侧面板用鼠标拖选选中当前2级码对应的1级码文本。
        前提：查询执行完毕，当前2级码已自动选中，右侧面板显示其1级码。
        
        校准点:
          right_panel_bottom — 「共X个追溯码」文字位置
          right_panel_top   — 右侧面板顶部第一个1级码上方空白处
        
        根据两个点的y坐标差计算拖动距离，适配任意窗口大小。
        
        返回: 纯1级码列表（已去除"复制""共X个追溯码"等杂乱信息）
        """
        bottom = self.cal.get("right_panel_bottom")
        top = self.cal.get("right_panel_top")
        if not bottom or not top:
            logger.warning("未校准 right_panel_bottom + right_panel_top，无法拖选")
            return []
        
        drag_y = bottom['y'] - top['y'] + 20  # y坐标差 + 20px余量
        drag_x = bottom['x'] - top['x'] + 20   # x坐标差 + 20px余量
        
        self._activate_window()
        # 点击「共X个追溯码」位置
        pyautogui.click(bottom['x'], bottom['y'])
        time.sleep(0.15)
        # 按住左键，向左上角拖动（按校准点计算的距离）
        pyautogui.mouseDown(button='left')
        time.sleep(0.05)
        pyautogui.move(-max(drag_x, 100), -max(drag_y, 100), duration=0.4)
        time.sleep(0.05)
        pyautogui.mouseUp(button='left')
        time.sleep(0.2)
        # 复制选中内容
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(0.35)
        
        raw = pyperclip.paste()
        logger.debug(f"拖选原始内容前200字: {raw[:200]}")
        
        # 清理杂乱信息：去掉"复制"、"共X个追溯码"等非码文本
        cleaned = []
        for line in raw.strip().split("\n"):
            line = line.strip()
            if not line:
                continue
            # 跳过"共X个追溯码"类文本
            if '共' in line and '个' in line and '追溯码' in line:
                continue
            # 去掉行尾的"复制"二字
            if line.endswith('复制'):
                line = line[:-2].strip()
            # 取纯数字码（16-24位纯数字）
            m = re.search(r'\b(\d{16,24})\b', line)
            if m:
                cleaned.append(m.group(1))
        
        logger.debug(f"拖选清理后获 {len(cleaned)} 个1级码")
        return cleaned

    def _handle_fail(self, code, result):
        """处理查询失败：通过回调弹窗让用户选择 继续/重试/停止"""
        if self.fail_callback:
            action = self.fail_callback(code, result.get("error", "未知错误"))
            if action == "retry":
                return "retry"
            elif action == "stop":
                raise StopIteration("用户选择停止")
            else:  # continue
                return "skip"
        return "skip"

    def batch_query_unified(self, records, input_trace_codes=None, callback=None):
        """
        统一批量查询（不分全1级/混装模式）。
        
        核心逻辑：
        - 用 covered_set 记录所有已查过的码（3级+2级+1级），排重
        - 用 output_1level_set 记录已输出的1级码，防重复
        - input_trace_codes: 输入表所有追溯码集合，供1级码同箱匹配
        - 先按码包装数降序排列，保证父码先处理
        - 一次查询两用：全量码→排重，1级码→输出
        """
        level_one_map = {}
        batch_no_map = {}
        covered_set = set()      # 所有已覆盖的码（3级+2级+1级）
        output_1level_set = set() # 已输出的1级码
        total = len(records)
        
        # 按码包装数降序排列（3级→2级→1级），保证父级先处理
        def sort_key(r):
            pc = r.code_pack_count
            if pc == "" or pc == "?":
                return 2  # 异常码放中间处理
            try:
                n = int(pc)
                if n > 1:
                    return 0  # 3级/2级码优先
                else:
                    return 1  # 1级码
            except:
                return 2
        
        sorted_records = sorted(records, key=sort_key)
        code_to_orig = {r.trace_code: r for r in records}
        
        for idx, record in enumerate(sorted_records, 1):
            if self.stop_flag.is_set():
                raise StopIteration("用户停止执行")
            
            code = record.trace_code
            pack_count = record.code_pack_count
            
            # ── 排重判断 ──
            if code in covered_set:
                if callback:
                    callback(idx, total, code, {"skip": True, "reason": f"已被其他码覆盖"})
                level_one_map[code] = []
                continue
            
            if pack_count == "1" and code in output_1level_set:
                if callback:
                    callback(idx, total, code, {"skip": True, "reason": "1级码已输出过"})
                level_one_map[code] = []
                continue
            
            # ── 开始查询 ──
            max_retries = 3
            for attempt in range(1, max_retries + 1):
                if self.stop_flag.is_set():
                    raise StopIteration("用户停止执行")
                
                try:
                    if pack_count == "1":
                        # 1级码：完整查询 → 查出同箱全部1级码 → 只保留输入表中也存在的（零散货匹配）
                        # 关键：只把实际输出的码加入covered_set，不覆盖整箱（防止混装不同批号被误跳）
                        if callback:
                            callback(idx, total, code, None)
                        result = self.query_single(code)
                        if result["success"]:
                            level_one = result.get("level_one_codes", [])
                            # 与输入表对比，只保留重复项（即零散货自身）
                            if input_trace_codes:
                                matched = [c for c in level_one if c in input_trace_codes]
                            else:
                                matched = level_one
                            new_codes = [c for c in matched if c not in output_1level_set]
                            if new_codes:
                                level_one_map[code] = new_codes
                                if result.get("batch_no"):
                                    batch_no_map[code] = result["batch_no"]
                                # 只加实际输出的，不加整箱（防止混装不同批号被误跳）
                                output_1level_set.update(new_codes)
                                covered_set.update(new_codes)
                            else:
                                level_one_map[code] = []
                            if callback:
                                callback(idx, total, code, result)
                            break
                        else:
                            if attempt < max_retries:
                                logger.debug(f"1级码查询失败，重试 ({attempt}/{max_retries})")
                                time.sleep(1)
                                continue
                            action = self._handle_fail(code, result)
                            if action == "retry":
                                attempt = 0
                                continue
                            elif action == "skip":
                                level_one_map[code] = [code]
                                if callback:
                                    callback(idx, total, code, result)
                                break
                    else:
                        # 2级/3级码：拖选精确模式 / 旧行为回退
                        if callback:
                            callback(idx, total, code, None)
                        result = self.query_single(code)
                        if result["success"]:
                            # 判断是否启用拖选精确模式
                            right_bottom = self.cal.get("right_panel_bottom")
                            right_top = self.cal.get("right_panel_top")
                            if right_bottom and right_top:
                                # 精确模式：拖选右侧面板获取当前2级码的精确1级码
                                exact_level_one = self._drag_select_right_panel()
                                level_one = exact_level_one if exact_level_one else result.get("level_one_codes", [])
                                new_codes = [c for c in level_one if c not in output_1level_set]
                                if new_codes:
                                    level_one_map[code] = new_codes
                                    if result.get("batch_no"):
                                        batch_no_map[code] = result["batch_no"]
                                    output_1level_set.update(new_codes)
                                    covered_set.update(new_codes)
                                else:
                                    level_one_map[code] = []
                                # 不覆盖整箱兄弟码
                                covered_set.add(code)
                            else:
                                # 回退模式：旧行为，全部码覆盖排重
                                all_codes = result.get("all_codes", set())
                                covered_set.update(all_codes)
                                level_one = result.get("level_one_codes", [])
                                new_codes = [c for c in level_one if c not in output_1level_set]
                                if new_codes:
                                    level_one_map[code] = new_codes
                                    if result.get("batch_no"):
                                        batch_no_map[code] = result["batch_no"]
                                    output_1level_set.update(new_codes)
                                else:
                                    level_one_map[code] = []
                            if callback:
                                callback(idx, total, code, result)
                            break
                        else:
                            if attempt < max_retries:
                                logger.debug(f"完整查询失败，重试 ({attempt}/{max_retries})")
                                time.sleep(1)
                                continue
                            # 全部重试失败
                            action = self._handle_fail(code, result)
                            if action == "retry":
                                attempt = 0
                                continue
                            elif action == "skip":
                                level_one_map[code] = [code]
                                if callback:
                                    callback(idx, total, code, result)
                                break
                except StopIteration:
                    raise
                except Exception as e:
                    logger.error(f"处理追溯码 {code} 出错: {str(e)}")
                    if attempt < max_retries:
                        time.sleep(1)
                        continue
                    action = self._handle_fail(code, {"error": str(e)})
                    if action == "retry":
                        attempt = 0
                        continue
                    elif action == "skip":
                        level_one_map[code] = [code]
                        break
                    else:
                        raise StopIteration("用户选择停止")
        
        return level_one_map, batch_no_map

# ============================================================
# 校准窗口
# ============================================================

class CalibrateWindow:
    """校准窗口 - 实时显示鼠标位置，用户按F5记录"""

    def __init__(self, parent, calibrator, on_complete):
        self.cal = calibrator
        self.on_complete = on_complete
        self.steps = [
            ("input_box", "请点击【追溯码输入框】"),
            ("query_btn", "请点击【查询】按钮"),
            ("copy_btn", "请点击【一键复制所有码】按钮（首次查询前位置）"),
            ("copy_btn_after", "请点击【一键复制所有码】按钮（查询结果后的位置）"),
            ("batch_no_pos", "请点击【产品批号】文字位置（用于复制批号）"),
            ("right_panel_bottom", "请点击【右侧面板】「共X个追溯码」文字位置"),
            ("right_panel_top", "请点击【右侧面板】顶部第一个1级码上方空白处"),
        ]
        self.current_step = 0
        self.current_x, self.current_y = 0, 0
        self.running = True
        self._f5_pending = False
        self._pynput_listener = None
        self._last_capture_time = 0  # F5防抖：防止多个监听重复触发

        self.win = tk.Toplevel(parent)
        self.win.title("校准 - 按F5记录位置")
        self.win.geometry("480x260")
        self.win.resizable(False, False)
        self.win.transient(parent)
        self.win.attributes('-topmost', True)

        # 绑定F5热键（窗口聚焦时生效）
        self.win.bind('<F5>', lambda e: self._capture_position())

        tk.Label(self.win, text="校准模式 - 按 F5 记录鼠标位置", font=("Microsoft YaHei", 13, "bold")).pack(pady=8)

        self.hint_label = tk.Label(self.win, text="", font=("Microsoft YaHei", 12), fg="#667eea")
        self.hint_label.pack(pady=4)

        coord_frame = tk.Frame(self.win, bg="#f0f0f0", padx=20, pady=10)
        coord_frame.pack(fill=tk.X, padx=20, pady=6)
        self.coord_label = tk.Label(coord_frame, text="当前鼠标位置: (0, 0)",
                                    font=("Consolas", 15), bg="#f0f0f0")
        self.coord_label.pack()

        self.pos_label = tk.Label(self.win, text="", font=("Microsoft YaHei", 10), fg="gray")
        self.pos_label.pack(pady=3)

        btn_frame = tk.Frame(self.win)
        btn_frame.pack(pady=6)

        tk.Button(btn_frame, text="按 F5 记录位置", font=("Microsoft YaHei", 11, "bold"),
                   bg="#667eea", fg="white", width=15, command=self._capture_position).pack(side=tk.LEFT, padx=10)

        tk.Button(btn_frame, text="测试点击", font=("Microsoft YaHei", 10),
                   bg="#52c41a", fg="white", width=10, command=self._test_click).pack(side=tk.LEFT, padx=10)

        # 提示
        tk.Label(self.win, text="F5 全局生效（切换到码上放心窗口也能按）",
                font=("Microsoft YaHei", 9),
                fg="#52c41a").pack(pady=3)

        # 启动鼠标跟踪线程
        self.mouse_thread = threading.Thread(target=self._track_mouse, daemon=True)
        self.mouse_thread.start()

        # ===== 全局F5监听：Windows API + pynput 双重保险 =====
        if PYNPUT_AVAILABLE:
            try:
                self._pynput_listener = pynput_keyboard.Listener(
                    on_press=self._on_global_key_press
                )
                self._pynput_listener.daemon = True
                self._pynput_listener.start()
            except Exception as e:
                print(f"[校准] pynput全局F5启动失败: {e}")

        # Windows API GetAsyncKeyState 轮询作为后备（最可靠）
        self._api_f5_thread = threading.Thread(target=self._api_f5_monitor, daemon=True)
        self._api_f5_thread.start()

        self._show_step()

        self.win.protocol("WM_DELETE_WINDOW", self._on_close)

    def _api_f5_monitor(self):
        """Windows API 轮询 F5 按键（全局生效，无需窗口焦点）"""
        import ctypes
        VK_F5 = 0x74
        last_state = False
        while self.running:
            try:
                current_state = bool(ctypes.windll.user32.GetAsyncKeyState(VK_F5) & 0x8000)
                if current_state and not last_state:
                    if self.win.winfo_exists():
                        self.win.after(0, self._capture_position)
                last_state = current_state
                time.sleep(0.1)
            except:
                time.sleep(0.1)

    def _on_global_key_press(self, key):
        """pynput全局键盘事件 - F5即使在码上放心窗口也能触发"""
        try:
            if key == pynput_keyboard.Key.f5:
                # 通过 after 切回主线程更新 UI
                if self.running and self.win.winfo_exists():
                    self.win.after(0, self._capture_position)
        except Exception:
            pass

    def _on_close(self):
        """窗口关闭时清理资源"""
        self.running = False
        if self._pynput_listener:
            try:
                self._pynput_listener.stop()
            except:
                pass
        try:
            self.win.destroy()
        except:
            pass
        # 触发完成回调
        try:
            self.on_complete()
        except:
            pass
    
    def _track_mouse(self):
        """后台线程：实时跟踪鼠标位置"""
        while self.running:
            try:
                x, y = pyautogui.position()
                self.current_x, self.current_y = x, y
                # 用默认参数固定闭包值，避免lambda引用变量
                self.win.after(0, lambda px=x, py=y: self.coord_label.config(
                    text=f"当前鼠标位置: ({px}, {py})"))
            except:
                pass
            time.sleep(0.1)  # 降低频率，减少事件队列压力
    
    def _show_step(self):
        if self.current_step >= len(self.steps):
            self.running = False
            self.win.destroy()
            self.on_complete()
            return
        name, hint = self.steps[self.current_step]
        self.hint_label.config(text=f"第{self.current_step+1}步: {hint}")
        recorded = self.cal.get(name)
        self.pos_label.config(text=f"已记录: ({recorded['x']}, {recorded['y']})" if recorded else "已记录: 未记录")
    
    def _capture_position(self):
        # 防抖：500ms内只响应一次F5
        now = time.time()
        if now - self._last_capture_time < 0.5:
            return
        self._last_capture_time = now
        
        name, _ = self.steps[self.current_step]
        x, y = self.current_x, self.current_y
        self.cal.set(name, x, y)
        self.pos_label.config(text=f"已记录: ({x}, {y}) ✓")
        self.current_step += 1
        self.win.update()  # 刷新界面
        self._show_step()
    
    def _test_click(self):
        name, _ = self.steps[self.current_step]
        pos = self.cal.get(name)
        if pos:
            pyautogui.click(pos['x'], pos['y'])
        else:
            messagebox.showinfo("提示", "请先按 F5 记录当前位置")

# ============================================================
# 主应用
# ============================================================
class DrugTraceApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"药品追溯码处理工具 {VERSION}")
        self.root.geometry("580x640")
        self.root.resizable(False, False)

        self.config = self._load_config()
        self.calibrator = ScreenCalibrator(os.path.join(DATA_DIR, 'calibration.json'))
        self.input_file = ""
        self.is_running = False
        self.stop_flag = threading.Event()
        self._f1_listener = None  # F1双击停止监听器
        self._f1_press_times = []  # F1按下时间记录
        self._create_widgets()
        self._update_calib_status()
    
    def _load_config(self):
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {'output_path': OUTPUT_FOLDER, 'input_path': DEFAULT_INPUT_PATH}
    
    def _save_config(self):
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(self.config, f, ensure_ascii=False, indent=2)
    
    def _update_calib_status(self):
        calibrated = self.calibrator.is_calibrated()
        if calibrated:
            self.calib_status.config(text="✅ 已校准", fg="#52c41a")
            self.start_btn.config(state=tk.NORMAL if self.input_file else tk.DISABLED)
        else:
            self.calib_status.config(text="❌ 未校准，请先校准", fg="#ff4d4f")
            self.start_btn.config(state=tk.DISABLED)
    
    def _start_emergency_stop(self):
        """后台监听 F1+F1 紧急停止（全局有效，无需焦点）"""
        VK_F1 = 0x70
        last_press = [0]  # 用列表避免闭包问题
        
        def monitor():
            import ctypes
            while True:
                if ctypes.windll.user32.GetAsyncKeyState(VK_F1) & 0x8000:
                    now = time.time()
                    if now - last_press[0] < 1.0:
                        # 两次 F1 间隔小于 1 秒，触发急停
                        self.stop_flag.set()
                        self.root.after(0, lambda: self.status_var.set("⚠️ 已紧急停止！"))
                        self.root.after(0, lambda: messagebox.showwarning("紧急停止", 
                            "已检测到 F1+F1 紧急停止指令，操作已中断。"))
                        last_press[0] = 0
                        time.sleep(1)  # 防重复触发
                    else:
                        last_press[0] = now
                    time.sleep(0.3)  # 消抖
                time.sleep(0.05)
        
        t = threading.Thread(target=monitor, daemon=True)
        t.start()
    
    def _create_widgets(self):
        # 配色方案（Win7-Win11 全兼容）
        BG_COLOR = "#F3F4F6"
        CARD_BG = "#FFFFFF"
        CARD_GLASS = "#FFFFFF"
        PRIMARY = "#4F46E5"
        PRIMARY_HOVER = "#4338CA"
        ACCENT_RED = "#EF4444"
        ACCENT_AMBER = "#F59E0B"
        TEXT_PRIMARY = "#1E293B"
        TEXT_SECONDARY = "#64748B"
        TEXT_MUTED = "#94A3B8"
        BORDER = "#E5E7EB"

        self.root.configure(bg=BG_COLOR)

        # ===== 标题区域 =====
        title_frame = tk.Frame(self.root, bg="#4F46E5", pady=14)
        title_frame.pack(fill=tk.X)

        # 标题下方的淡色装饰条
        accent_bar = tk.Frame(self.root, bg="#818CF8", height=2)
        accent_bar.pack(fill=tk.X)

        title_inner = tk.Frame(title_frame, bg="#4F46E5")
        title_inner.pack()
        tk.Label(title_inner, text="💊  药品批发企业追朔码自动处理软件",
                font=("Microsoft YaHei", 14, "bold"), bg="#4F46E5", fg="white").pack(side=tk.LEFT)
        tk.Label(title_inner, text=f"  {VERSION}", font=("Microsoft YaHei", 8, "bold"),
                bg="#F97316", fg="white", padx=5, pady=2).pack(side=tk.LEFT, padx=(5, 0))
        tk.Label(title_frame, text="自动查询追溯码关联关系，生成1级码表格  |  紧急停止: 双击 F1",
                font=("Microsoft YaHei", 8), bg="#4F46E5", fg="#C7D2FE").pack(pady=(2, 0))

        # ===== 主区域 =====
        main_frame = tk.Frame(self.root, bg=BG_COLOR, padx=24, pady=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # ---- 卡片辅助函数（左侧装饰色条）----
        def make_glass_card(parent, title_icon_text, accent_color="#4F46E5"):
            outer = tk.Frame(parent, bg=CARD_GLASS, highlightbackground=BORDER,
                            highlightthickness=1)
            outer.pack(fill=tk.X, pady=(0, 8))

            # 左侧色条装饰
            stripe = tk.Frame(outer, bg=accent_color, width=4)
            stripe.pack(side=tk.LEFT, fill=tk.Y)
            stripe.pack_propagate(False)

            # 内容区
            body = tk.Frame(outer, bg=CARD_GLASS, padx=14, pady=10)
            body.pack(side=tk.LEFT, fill=tk.X, expand=True)

            # 标题行
            hdr = tk.Frame(body, bg=CARD_GLASS)
            hdr.pack(fill=tk.X, pady=(0, 6))
            tk.Label(hdr, text=title_icon_text, font=("Microsoft YaHei", 10, "bold"),
                    bg=CARD_GLASS, fg=TEXT_PRIMARY).pack(side=tk.LEFT)
            return outer, body, hdr

        # ===== 卡片1：校准（amber 色条）=====
        card1, body1, hdr1 = make_glass_card(main_frame, "🔧 按钮校准", ACCENT_AMBER)
        self.calib_status = tk.Label(hdr1, text="", font=("Microsoft YaHei", 8),
                                     bg=CARD_GLASS)
        self.calib_status.pack(side=tk.LEFT, padx=8)
        tk.Button(hdr1, text="校准位置", font=("Microsoft YaHei", 9, "bold"),
                 bg=ACCENT_AMBER, fg="white", relief=tk.FLAT, padx=14, pady=3,
                 command=self._start_calibrate, cursor="hand2",
                 activebackground="#D97706").pack(side=tk.RIGHT)

        # ===== 卡片2：文件选择（indigo 色条）=====
        card2, body2, _ = make_glass_card(main_frame, "📂 文件选择", "#4F46E5")

        # 输入文件
        in_row = tk.Frame(body2, bg=CARD_GLASS)
        in_row.pack(fill=tk.X, pady=(0, 5))
        tk.Label(in_row, text="输入文件", font=("Microsoft YaHei", 9),
                bg=CARD_GLASS, fg=TEXT_PRIMARY, width=7, anchor="w").pack(side=tk.LEFT)
        self.file_var = tk.StringVar(value="点击右侧按钮选择文件...")
        tk.Entry(in_row, textvariable=self.file_var, state="readonly",
                font=("Microsoft YaHei", 9), relief=tk.FLAT, bg="#F8FAFC", bd=1,
                highlightbackground="#E5E7EB", highlightthickness=1).pack(
                    side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6), ipady=4)
        tk.Button(in_row, text="浏览", command=self._select_file,
                 font=("Microsoft YaHei", 9), width=6, relief=tk.FLAT,
                 bg="#F1F5F9", fg=TEXT_PRIMARY, activebackground="#E2E8F0",
                 cursor="hand2", padx=4, pady=2).pack(side=tk.RIGHT)

        # 导出位置
        out_row = tk.Frame(body2, bg=CARD_GLASS)
        out_row.pack(fill=tk.X, pady=(0, 5))
        tk.Label(out_row, text="导出位置", font=("Microsoft YaHei", 9),
                bg=CARD_GLASS, fg=TEXT_PRIMARY, width=7, anchor="w").pack(side=tk.LEFT)
        self.out_var = tk.StringVar(value=self.config.get('output_path', OUTPUT_FOLDER))
        tk.Entry(out_row, textvariable=self.out_var, state="readonly",
                font=("Microsoft YaHei", 9), relief=tk.FLAT, bg="#F8FAFC", bd=1,
                highlightbackground="#E5E7EB", highlightthickness=1).pack(
                    side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6), ipady=4)
        tk.Button(out_row, text="浏览", command=self._select_output,
                 font=("Microsoft YaHei", 9), width=6, relief=tk.FLAT,
                 bg="#F1F5F9", fg=TEXT_PRIMARY, activebackground="#E2E8F0",
                 cursor="hand2", padx=4, pady=2).pack(side=tk.RIGHT)

        # 分隔线 + 默认路径
        sep = ttk.Separator(body2, orient='horizontal')
        sep.pack(fill=tk.X, pady=(4, 6))
        path_hdr = tk.Frame(body2, bg=CARD_GLASS)
        path_hdr.pack(fill=tk.X)
        tk.Label(path_hdr, text="⚙️ 默认路径", font=("Microsoft YaHei", 8, "bold"),
                bg=CARD_GLASS, fg=TEXT_SECONDARY).pack(side=tk.LEFT)

        def_path_row = tk.Frame(body2, bg=CARD_GLASS)
        def_path_row.pack(fill=tk.X, pady=(4, 0))
        self.default_input_var = tk.StringVar(value=self.config.get('input_path', DEFAULT_INPUT_PATH))
        tk.Label(def_path_row, text="输入", font=("Microsoft YaHei", 8),
                bg=CARD_GLASS, fg=TEXT_MUTED, width=3, anchor="w").pack(side=tk.LEFT)
        tk.Entry(def_path_row, textvariable=self.default_input_var, state="readonly",
                font=("Microsoft YaHei", 8), relief=tk.FLAT, bg="#F8FAFC", bd=1,
                highlightbackground="#E5E7EB", highlightthickness=1).pack(
                    side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4), ipady=2)
        tk.Button(def_path_row, text="修改", command=self._set_default_input,
                 font=("Microsoft YaHei", 8), width=4, relief=tk.FLAT,
                 bg=PRIMARY, fg="white", activebackground=PRIMARY_HOVER,
                 cursor="hand2", padx=6, pady=1).pack(side=tk.RIGHT)

        def_path_row2 = tk.Frame(body2, bg=CARD_GLASS)
        def_path_row2.pack(fill=tk.X, pady=(3, 0))
        self.default_output_var = tk.StringVar(value=self.config.get('output_path', OUTPUT_FOLDER))
        tk.Label(def_path_row2, text="输出", font=("Microsoft YaHei", 8),
                bg=CARD_GLASS, fg=TEXT_MUTED, width=3, anchor="w").pack(side=tk.LEFT)
        tk.Entry(def_path_row2, textvariable=self.default_output_var, state="readonly",
                font=("Microsoft YaHei", 8), relief=tk.FLAT, bg="#F8FAFC", bd=1,
                highlightbackground="#E5E7EB", highlightthickness=1).pack(
                    side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4), ipady=2)
        tk.Button(def_path_row2, text="修改", command=self._set_default_output,
                 font=("Microsoft YaHei", 8), width=4, relief=tk.FLAT,
                 bg=PRIMARY, fg="white", activebackground=PRIMARY_HOVER,
                 cursor="hand2", padx=6, pady=1).pack(side=tk.RIGHT)

        # ===== 卡片3：单位名称设置（teal 色条）=====
        card3, body3, _ = make_glass_card(main_frame, "🏢 单位名称设置", "#14B8A6")

        unit_row = tk.Frame(body3, bg=CARD_GLASS)
        unit_row.pack(fill=tk.X)
        self.unit_name_var = tk.StringVar(value=self.config.get('unit_name', '药品批发企业'))
        tk.Entry(unit_row, textvariable=self.unit_name_var,
                font=("Microsoft YaHei", 9), relief=tk.SOLID, bd=1,
                bg="#F8FAFC", highlightbackground="#E5E7EB", highlightthickness=1).pack(
                    side=tk.LEFT, fill=tk.X, expand=True, ipady=4)
        tk.Label(unit_row, text=" 生成文件名前缀",
                font=("Microsoft YaHei", 8), bg=CARD_GLASS, fg=TEXT_MUTED).pack(side=tk.LEFT, padx=(5, 0))

        # ===== 操作按钮 =====
        btn_frame = tk.Frame(main_frame, bg=BG_COLOR)
        btn_frame.pack(fill=tk.X, pady=(4, 6))

        self.start_btn = tk.Button(btn_frame, text="  ▶   开始处理  ", command=self._start_process,
                                   font=("Microsoft YaHei", 11, "bold"), bg=PRIMARY, fg="white",
                                   height=1, relief=tk.FLAT, cursor="hand2",
                                   activebackground=PRIMARY_HOVER, padx=16, pady=7,
                                   disabledforeground="#A5B4FC")  # 禁用时文字仍可见
        self.start_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))

        self.stop_btn = tk.Button(btn_frame, text="  ⏹   停止  ", command=self._stop_process,
                                  font=("Microsoft YaHei", 11, "bold"), bg=ACCENT_RED, fg="white",
                                  height=1, relief=tk.FLAT, cursor="hand2",
                                  activebackground="#DC2626", padx=16, pady=7,
                                  disabledforeground="#FCA5A5")
        self.stop_btn.pack(side=tk.RIGHT, ipadx=24)

        # ===== 进度条 =====
        style = ttk.Style()
        style.theme_use('default')
        style.configure("Custom.Horizontal.TProgressbar", troughcolor='#E2E8F0',
                        background="#4F46E5", thickness=8, borderwidth=0)
        self.progress_var = tk.DoubleVar(value=0)
        ttk.Progressbar(main_frame, variable=self.progress_var, maximum=100,
                        style="Custom.Horizontal.TProgressbar").pack(fill=tk.X, pady=(0, 4))
        self.status_var = tk.StringVar(value="请先校准按钮位置，然后选择Excel文件")
        tk.Label(main_frame, textvariable=self.status_var, font=("Microsoft YaHei", 9),
                fg=TEXT_SECONDARY, bg=BG_COLOR).pack(fill=tk.X)

        # ===== 底部提示 =====
        hint_frame = tk.Frame(main_frame, bg="#F0FDF4", highlightbackground="#BBF7D0",
                              highlightthickness=1)
        hint_frame.pack(fill=tk.X)
        tk.Label(hint_frame,
                text=" ① 校准 ② 选Excel ③ 开始处理（自动排重）  |  ⛔ 紧急停止: 双击F1",
                font=("Microsoft YaHei", 8), bg="#F0FDF4", fg="#166534",
                padx=8, pady=5).pack()
    
    def _start_calibrate(self):
        CalibrateWindow(self.root, self.calibrator, self._on_calibrate_complete)
    
    def _on_calibrate_complete(self):
        self._update_calib_status()
        messagebox.showinfo("校准完成", "按钮位置已记录！\n\n注意：如果码上放心窗口移动了，需要重新校准。")
    
    def _select_file(self):
        default_dir = self.config.get('input_path', DEFAULT_INPUT_PATH)
        if not os.path.exists(default_dir):
            default_dir = os.path.expanduser("~\\Desktop")
        file_path = filedialog.askopenfilename(title="选择Excel文件",
                                              initialdir=default_dir,
                                              filetypes=[("Excel文件", "*.xlsx")])
        if file_path:
            self.input_file = file_path
            self.file_var.set(file_path)
            if self.calibrator.is_calibrated():
                self.start_btn.config(state=tk.NORMAL)
            self.status_var.set("文件已选择，点击开始处理")
    
    def _select_output(self):
        file_path = filedialog.askopenfilename(title="请选择一个文件，程序将把输出文件保存在同目录下",
                                              filetypes=[("所有文件", "*.*")])
        if file_path:
            folder = os.path.dirname(file_path)
            self.config['output_path'] = folder
            self._save_config()
            self.out_var.set(folder)
            self.default_output_var.set(folder)
    
    def _set_default_input(self):
        folder = filedialog.askdirectory(title="选择默认输入目录",
                                         initialdir=self.config.get('input_path', DEFAULT_INPUT_PATH))
        if folder:
            self.config['input_path'] = folder
            self._save_config()
            self.default_input_var.set(folder)
            self.status_var.set(f"默认输入路径已设置为: {folder}")
    
    def _set_default_output(self):
        folder = filedialog.askdirectory(title="选择默认输出目录",
                                         initialdir=self.config.get('output_path', OUTPUT_FOLDER))
        if folder:
            self.config['output_path'] = folder
            self._save_config()
            self.out_var.set(folder)
            self.default_output_var.set(folder)
            self.status_var.set(f"默认输出路径已设置为: {folder}")
    
    def _stop_process(self):
        if self.is_running:
            self.stop_flag.set()
            self.status_var.set("正在停止...")

    def _start_f1_listener(self):
        """启动全局F1双击监听（仅处理中生效）"""
        if not PYNPUT_AVAILABLE:
            return
        self._f1_press_times = []
        try:
            self._f1_listener = pynput_keyboard.Listener(
                on_press=self._on_f1_press
            )
            self._f1_listener.daemon = True
            self._f1_listener.start()
        except Exception as e:
            print(f"[F1监听] 启动失败: {e}")

    def _stop_f1_listener(self):
        """停止F1监听"""
        if self._f1_listener:
            try:
                self._f1_listener.stop()
            except:
                pass
            self._f1_listener = None

    def _on_f1_press(self, key):
        """全局F1按键事件 - 双击1.5秒内触发强制停止"""
        try:
            if key == pynput_keyboard.Key.f1 and self.is_running:
                now = time.time()
                # 清理超过1.5秒的旧记录
                self._f1_press_times = [t for t in self._f1_press_times if now - t < 1.5]
                self._f1_press_times.append(now)
                # 单次F1 = 提示，双击 = 强制停止
                if len(self._f1_press_times) >= 2:
                    self._f1_press_times = []
                    self.stop_flag.set()
                    self.root.after(0, lambda: self.status_var.set("⚠️ 检测到双击 F1，强制停止中..."))
                else:
                    self.root.after(0, lambda: self.status_var.set("⚠️ 再按一次 F1 确认停止"))
        except Exception:
            pass
    
    def _show_warning_dialog(self, total):
        """显示操作前警告对话框 - 居中显示"""
        win = tk.Toplevel(self.root)
        win.title("⚠️  操作确认")
        win_w, win_h = 440, 260
        win.resizable(False, False)
        win.transient(self.root)
        win.grab_set()
        win.attributes('-topmost', True)

        # 居中显示在主窗口中央
        self.root.update_idletasks()
        parent_x = self.root.winfo_x()
        parent_y = self.root.winfo_y()
        parent_w = self.root.winfo_width()
        parent_h = self.root.winfo_height()
        x = parent_x + (parent_w - win_w) // 2
        y = parent_y + (parent_h - win_h) // 2
        win.geometry(f"{win_w}x{win_h}+{x}+{y}")

        # 红底标题
        title_frame = tk.Frame(win, bg="#cf1322", pady=12)
        title_frame.pack(fill=tk.X)
        tk.Label(title_frame, text="⚠️  电脑即将自动操作",
                font=("Microsoft YaHei", 14, "bold"), bg="#cf1322", fg="white").pack()

        body_frame = tk.Frame(win, padx=24, pady=10)
        body_frame.pack(fill=tk.BOTH, expand=True)

        warning_lines = [
            f"自动排重模式  |  待处理: {total} 条",
            "",
            "请确保码上放心客户端已打开！",
            "操作期间请勿移动鼠标或敲击键盘！",
            "⚡ 紧急停止: 双击 F1 键",
        ]
        for line in warning_lines:
            fg_color = "#cf1322" if ("不要" in line or "确保" in line or "F1" in line) else "#333"
            tk.Label(body_frame, text=line, font=("Microsoft YaHei", 10, "bold" if fg_color=="#cf1322" else "normal"),
                    fg=fg_color).pack(anchor="w", pady=1)

        # 按钮
        btn_frame = tk.Frame(win, pady=8)
        btn_frame.pack(fill=tk.X)

        result_flag = {"confirmed": False}

        def on_confirm():
            result_flag["confirmed"] = True
            win.destroy()

        def on_cancel():
            win.destroy()

        tk.Button(btn_frame, text="确 定", command=on_confirm,
                 font=("Microsoft YaHei", 11, "bold"), bg="#cf1322", fg="white",
                 width=10, height=1, relief=tk.FLAT, cursor="hand2").pack(side=tk.LEFT, padx=(20, 10))
        tk.Button(btn_frame, text="取 消", command=on_cancel,
                 font=("Microsoft YaHei", 11), bg="#e8e8e8", fg="#333",
                 width=10, height=1, relief=tk.FLAT, cursor="hand2").pack(side=tk.RIGHT, padx=(10, 20))

        self.root.wait_window(win)
        return result_flag["confirmed"]
    
    def _show_fail_dialog(self, code, error_msg):
        """查询失败弹窗报警 - 用户选择 继续/重试/停止（居中，支持F1急停）"""
        result = {"action": "skip"}
        done = threading.Event()

        def show():
            dialog = tk.Toplevel(self.root)
            dialog.title("查询失败")
            dialog.transient(self.root)
            dialog.grab_set()
            dialog.attributes('-topmost', True)
            dialog.resizable(False, False)

            # 居中
            self.root.update_idletasks()
            px, py = self.root.winfo_x(), self.root.winfo_y()
            pw, ph = self.root.winfo_width(), self.root.winfo_height()
            dw, dh = 380, 160
            cx = px + (pw - dw) // 2 if pw > dw else (self.root.winfo_screenwidth() - dw) // 2
            cy = py + (ph - dh) // 2 if ph > dh else (self.root.winfo_screenheight() - dh) // 2
            dialog.geometry(f"{dw}x{dh}+{cx}+{cy}")

            title_f = tk.Frame(dialog, bg="#cf1322", pady=6)
            title_f.pack(fill=tk.X)
            tk.Label(title_f, text="查询失败", font=("Microsoft YaHei", 12, "bold"),
                    bg="#cf1322", fg="white").pack()

            body = tk.Frame(dialog, padx=16, pady=6)
            body.pack(fill=tk.BOTH, expand=True)
            tk.Label(body, text=f"追溯码: {code}", font=("Microsoft YaHei", 9),
                    fg="#333", anchor="w").pack(anchor="w")
            tk.Label(body, text=error_msg[:60], font=("Microsoft YaHei", 8),
                    fg="#999", anchor="w").pack(anchor="w", pady=(0, 4))
            tk.Label(body, text="F1+F1 可紧急停止", font=("Microsoft YaHei", 8),
                    fg="#cf1322", anchor="w").pack(anchor="w")

            btn_f = tk.Frame(dialog, pady=4)
            btn_f.pack(fill=tk.X)

            def choose(action):
                result["action"] = action
                dialog.destroy()
                done.set()

            tk.Button(btn_f, text="跳过", command=lambda: choose("skip"),
                     font=("Microsoft YaHei", 9), bg="#faad14", fg="white",
                     width=8, relief=tk.FLAT, cursor="hand2").pack(side=tk.LEFT, padx=(12, 4))
            tk.Button(btn_f, text="重试", command=lambda: choose("retry"),
                     font=("Microsoft YaHei", 9, "bold"), bg="#52c41a", fg="white",
                     width=6, relief=tk.FLAT, cursor="hand2").pack(side=tk.LEFT, padx=4)
            tk.Button(btn_f, text="停止", command=lambda: choose("stop"),
                     font=("Microsoft YaHei", 9), bg="#ff4d4f", fg="white",
                     width=6, relief=tk.FLAT, cursor="hand2").pack(side=tk.RIGHT, padx=(4, 12))

            dialog.wait_window()

        self.root.after(0, show)
        done.wait()  # 等待对话框关闭再返回
        return result["action"]

    def _start_process(self):
        if not self.calibrator.is_calibrated():
            messagebox.showerror("错误", "请先完成按钮位置校准")
            return
        if not self.input_file or not os.path.exists(self.input_file):
            messagebox.showerror("错误", "请选择有效的Excel文件")
            return
        if self.is_running:
            return
        
        # 保存单位名称到配置
        unit_name = self.unit_name_var.get().strip() or "药品批发企业"
        self.config['unit_name'] = unit_name
        self._save_config()
        
        # 统计总行数
        import openpyxl
        wb = openpyxl.load_workbook(self.input_file, data_only=True)
        ws = wb.active
        total_rows = 0
        for row_idx in range(7, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=1).value
            if val:
                total_rows += 1
        wb.close()
        
        # 显示警告对话框
        if not self._show_warning_dialog(total_rows):
            self.status_var.set("操作已取消")
            return
        
        # 1.5秒倒计时后自动开始
        self.status_var.set("⏳ 1.5 秒后自动激活码上放心操作...（请勿移动鼠标键盘）")
        time.sleep(1.5)
        
        self.stop_flag.clear()
        self.is_running = True
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.status_var.set("正在启动...（不要移动鼠标键盘）")
        # 启动F1双击全局监听（紧急停止用）
        self._start_f1_listener()
        Thread(target=self._process_file, daemon=True).start()
    
    def _process_file(self):
        try:
            unit_name = self.config.get('unit_name', '药品批发企业')
            
            # 1. 自动激活码上放心窗口
            self.root.after(0, lambda: self.status_var.set("正在激活码上放心客户端窗口..."))
            try:
                windows = pyautogui.getWindowsWithTitle("码上放心")
                if windows:
                    windows[0].activate()
                    time.sleep(0.5)
                else:
                    self.root.after(0, lambda: messagebox.showinfo("提示", "未找到码上放心窗口，请手动切换"))
                    time.sleep(3)
            except:
                time.sleep(3)
            
            # 2. 读取Excel
            self.root.after(0, lambda: self.status_var.set("正在读取Excel文件..."))
            self.root.after(0, lambda: self.progress_var.set(5))

            data = read_sales_excel(self.input_file)
            if not data.records:
                self.root.after(0, lambda: self.status_var.set("错误：未找到追溯码记录"))
                return

            # 3. 进度恢复检测（崩溃续跑）
            progress_file = os.path.join(DATA_DIR, '_progress.json')
            processed_indices = set()
            if os.path.exists(progress_file):
                try:
                    with open(progress_file, 'r') as f:
                        prog = json.load(f)
                    if prog.get('file') == self.input_file:
                        processed_indices = set(prog.get('processed', []))
                        self.root.after(0, lambda p=len(processed_indices): self.status_var.set(
                            f"检测到上次处理进度，已跳过 {p} 条已处理记录"))
                except:
                    pass

            # 过滤已处理的记录
            if processed_indices:
                data.records = [r for i, r in enumerate(data.records) if i not in processed_indices]
                if not data.records:
                    self.root.after(0, lambda: self.status_var.set("所有记录已处理完毕"))
                    return

            total = len(data.records)

            # 进度保存映射
            code_to_orig_idx = {r.trace_code: i for i, r in enumerate(data.records)}

            def save_progress(code):
                orig_idx = code_to_orig_idx.get(code)
                if orig_idx is not None:
                    processed_indices.add(orig_idx)
                try:
                    with open(progress_file, 'w', encoding='utf-8') as f:
                        json.dump({
                            'file': self.input_file,
                            'processed': sorted(processed_indices),
                        }, f, ensure_ascii=False)
                except Exception as e:
                    logger.debug(f"保存进度失败: {e}")

            self.root.after(0, lambda t=total: self.status_var.set(
                f"共{t}条追溯码，统一排重模式，开始处理..."))
            self.root.after(0, lambda: self.progress_var.set(10))

            # 建立失败回调（弹窗报警）
            def on_query_fail(code, error_msg):
                return self._show_fail_dialog(code, error_msg)

            ui = CodeQueryUI(self.calibrator, self.stop_flag, fail_callback=on_query_fail)

            def on_progress(idx, t, code, result):
                pct = 10 + (idx / t) * 80
                self.root.after(0, lambda p=pct: self.progress_var.set(p))
                if result:
                    if result.get('skip'):
                        reason = result.get('reason', '已处理')
                        self.root.after(0, lambda i=idx, tt=t, r=reason: self.status_var.set(f"处理中: {i}/{tt} - 跳过（{r}）"))
                    elif result.get('success'):
                        n = len(result.get('level_one_codes', []))
                        self.root.after(0, lambda i=idx, tt=t, nn=n: self.status_var.set(f"处理中: {i}/{tt} - 获取{nn}个1级码"))
                        save_progress(code)
                    else:
                        err = result.get('error', '失败')
                        self.root.after(0, lambda i=idx, tt=t, e=err: self.status_var.set(f"处理中: {i}/{tt} - {e}"))
                else:
                    self.root.after(0, lambda i=idx, tt=t: self.status_var.set(f"处理中: {i}/{tt}"))

            # 4. 执行统一批处理
            level_one_map = {}
            batch_no_map = {}

            # 进度恢复：先填已处理的行
            if processed_indices:
                for idx in processed_indices:
                    if idx < len(data.records):
                        r = data.records[idx]
                        level_one_map[r.trace_code] = [r.trace_code]
                        batch_no_map[r.trace_code] = r.batch_no

            # 过滤出未处理的行
            unprocessed = [r for i, r in enumerate(data.records) if i not in processed_indices]

            if not unprocessed:
                self.root.after(0, lambda: self.status_var.set("所有记录已处理完毕"))
                self.root.after(0, lambda: self.progress_var.set(90))
            else:
                # 构建输入表追溯码集合（供1级码同箱匹配）
                input_trace_set = set(r.trace_code for r in data.records)
                level_one_map, batch_no_map = ui.batch_query_unified(
                    unprocessed, input_trace_codes=input_trace_set, callback=on_progress)
            
            # 5. 生成Excel
            self.root.after(0, lambda: self.status_var.set("正在生成Excel文件..."))
            self.root.after(0, lambda: self.progress_var.set(95))
            output_file = write_result_excel(data, level_one_map, batch_no_map, self.out_var.get(), unit_name)
            
            # 6. 清理进度文件（处理成功）
            if os.path.exists(progress_file):
                os.remove(progress_file)
            
            self.root.after(0, lambda: self.progress_var.set(100))
            self.root.after(0, lambda: messagebox.showinfo("完成", f"处理完成！\n文件已保存到:\n{output_file}"))
            
        except StopIteration:
            self._on_process_stop()
            return
        except Exception as e:
            self.root.after(0, lambda err=str(e): self.status_var.set(f"错误: {err}"))
            self.root.after(0, lambda err=str(e): messagebox.showerror("错误", str(e)))
        finally:
            self._reset_ui()
    
    def _on_process_stop(self):
        self.root.after(0, lambda: self.status_var.set("处理已停止"))
        self.root.after(0, lambda: messagebox.showinfo("已停止", "处理已手动停止。\n可以重新开始处理，或修改后继续。"))
        self._reset_ui()
    
    def _reset_ui(self):
        self.is_running = False
        self.stop_flag.clear()
        # 停止F1监听
        self._stop_f1_listener()
        self.root.after(0, lambda: self.start_btn.config(state=tk.NORMAL, bg="#667eea"))
        self.root.after(0, lambda: self.stop_btn.config(state=tk.DISABLED))


def main():
    try:
        # 启动时检查更新
        if check_and_update():
            sys.exit(0)

        root = tk.Tk()
        DrugTraceApp(root)
        root.mainloop()
    except Exception as e:
        import traceback
        error_log = os.path.join(DATA_DIR, 'crash.log')
        with open(error_log, 'w', encoding='utf-8') as f:
            f.write(f"启动失败: {e}\n")
            traceback.print_exc(file=f)
        # 弹出错误提示
        try:
            import ctypes
            ctypes.windll.user32.MessageBoxW(0, f"启动失败！\n\n错误: {e}\n\n详情请查看:\n{error_log}", "错误", 0x10)
        except:
            pass

if __name__ == '__main__':
    main()
